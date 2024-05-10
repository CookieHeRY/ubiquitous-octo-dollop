# -*- coding: utf-8 -*-
# /usr/bin/env/python3

'''
Pytorch implementation for LPRNet.
Author: aiboy.wei@outlook.com .
'''

from data.load_data import CHARS, CHARS_DICT, LPRDataLoader
from model.LPRNet import build_lprnet
# import torch.backends.cudnn as cudnn
from torch.autograd import Variable
import torch.nn.functional as F
from torch.utils.data import *
from torch import optim
import torch.nn as nn
import numpy as np
import argparse
import torch
import time
import os
import os.path
from openpyxl import load_workbook
import pandas as pd

excel_name = r"results.xlsx"
sheet_name = "Sheet1"


def sparse_tuple_for_ctc(T_length, lengths):
    input_lengths = []
    target_lengths = []

    for ch in lengths:
        input_lengths.append(T_length)
        target_lengths.append(ch)

    return tuple(input_lengths), tuple(target_lengths)


def adjust_learning_rate(optimizer, cur_epoch, base_lr, lr_schedule):
    """
    Sets the learning rate
    """
    lr = 0
    for i, e in enumerate(lr_schedule):
        if cur_epoch < e:
            lr = base_lr * (0.1 ** i)
            break
    if lr == 0:
        lr = base_lr
    for param_group in optimizer.param_groups:
        param_group['lr'] = lr

    return lr


def write_excel(excel_name, sheet_name, value):
    columns = ["epoc", "train_loss", "test_acc"]  # 列名
    # 创建一个 pandas 的数据框
    if not os.path.exists(excel_name):  # 文件不存在,就创建一个
        df = pd.DataFrame(columns=columns)
        df.to_excel(excel_name, index=False)
    books = load_workbook(excel_name).sheetnames  # 得到已存在的sheet列表

    # 将数据框写入 Excel 文件
    if sheet_name not in books:  # 如果sheet_name不存在,创建
        with pd.ExcelWriter(excel_name, engine="openpyxl", mode="a") as writer:
            df = pd.DataFrame(columns=columns)
            df.to_excel(writer, sheet_name=sheet_name, index=False)  # header=None 参数用于追加写入时不重复写入列名
    # 追加一行数据
    workbooks = load_workbook(excel_name)  # 获取文件
    worksheet = workbooks[sheet_name]  # 获取工作表sheet
    worksheet._current_row = worksheet.max_row  # 指定最大行数
    worksheet.append(value)  # 添加数据
    workbooks.save(excel_name)  # 保存


def get_parser():
    parser = argparse.ArgumentParser(description='parameters to train net')
    parser.add_argument('--max_epoch', default=100, help='epoch to train the network')
    parser.add_argument('--img_size', default=[94, 24], help='the image size')
    parser.add_argument('--train_img_dirs', default=r"F:\machinelearning\project\datasets\ccpd\lpr3\train",
                        help='the train images path')
    parser.add_argument('--test_img_dirs', default=r"F:\machinelearning\project\datasets\ccpd\lpr3\val",
                        help='the test images path')
    parser.add_argument('--dropout_rate', default=0.5, help='dropout rate.')
    parser.add_argument('--learning_rate', default=0.01, help='base value of learning rate.')
    parser.add_argument('--lpr_max_len', default=8, help='license plate number max length.')
    parser.add_argument('--train_batch_size', default=128, help='training batch size.')
    parser.add_argument('--test_batch_size', default=128, help='testing batch size.')
    parser.add_argument('--phase_train', default=True, type=bool, help='train or test phase flag.')
    parser.add_argument('--num_workers', default=0, type=int, help='Number of workers used in dataloading')
    parser.add_argument('--cuda', default=True, type=bool, help='Use cuda to train model')
    parser.add_argument('--resume_epoch', default=0, type=int, help='resume iter for retraining')
    parser.add_argument('--save_interval', default=2000, type=int, help='interval for save model state dict')
    parser.add_argument('--test_interval', default=2000, type=int, help='interval for evaluate')
    parser.add_argument('--momentum', default=0.9, type=float, help='momentum')
    parser.add_argument('--weight_decay', default=2e-5, type=float, help='Weight decay for SGD')
    parser.add_argument('--lr_schedule', default=[20, 40, 60, 80, 100], help='schedule for learning rate.')
    parser.add_argument('--save_folder', default='./weights/', help='Location to save checkpoint models')
    parser.add_argument('--show', default=False, help='show the result')
    # parser.add_argument('--pretrained_model', default='../weights/LPRNet__0.795.pth', help='pretrained base model')
    parser.add_argument('--pretrained_model', default='../weights/Final_LPRNet_model.pth', help='pretrained base model')

    args = parser.parse_args()

    return args


def collate_fn(batch):
    imgs = []
    labels = []
    lengths = []
    for _, sample in enumerate(batch):
        img, label, length = sample
        imgs.append(torch.from_numpy(img))
        labels.extend(label)
        lengths.append(length)
    labels = np.asarray(labels).flatten().astype(int)

    return (torch.stack(imgs, 0), torch.from_numpy(labels), lengths)


def train():
    args = get_parser()
    bestAcc = 0

    T_length = 18  # args.lpr_max_len
    epoch = 0 + args.resume_epoch
    loss_val = 0

    if not os.path.exists(args.save_folder):
        os.mkdir(args.save_folder)

    lprnet = build_lprnet(lpr_max_len=args.lpr_max_len, phase=args.phase_train, class_num=len(CHARS),
                          dropout_rate=args.dropout_rate)
    device = torch.device("cuda:0" if args.cuda else "cpu")
    lprnet.to(device)
    print("Successful to build network!")

    # load pretrained model
    if args.pretrained_model:
        lprnet.load_state_dict(torch.load(args.pretrained_model))
        print("load pretrained model successful!")
    else:
        def xavier(param):
            nn.init.xavier_uniform(param)

        def weights_init(m):
            for key in m.state_dict():
                if key.split('.')[-1] == 'weight':
                    if 'conv' in key:
                        nn.init.kaiming_normal_(m.state_dict()[key], mode='fan_out')
                    if 'bn' in key:
                        m.state_dict()[key][...] = xavier(1)
                elif key.split('.')[-1] == 'bias':
                    m.state_dict()[key][...] = 0.01

        lprnet.backbone.apply(weights_init)
        lprnet.container.apply(weights_init)
        print("initial net weights successful!")

    # define optimizer
    # optimizer = optim.SGD(lprnet.parameters(), lr=args.learning_rate,
    #                       momentum=args.momentum, weight_decay=args.weight_decay)
    optimizer = optim.RMSprop(lprnet.parameters(), lr=args.learning_rate, alpha=0.9, eps=1e-08,
                              momentum=args.momentum, weight_decay=args.weight_decay)
    train_img_dirs = os.path.expanduser(args.train_img_dirs)
    test_img_dirs = os.path.expanduser(args.test_img_dirs)
    train_dataset = LPRDataLoader(train_img_dirs.split(','), args.img_size, args.lpr_max_len)
    test_dataset = LPRDataLoader(test_img_dirs.split(','), args.img_size, args.lpr_max_len)

    epoch_size = len(train_dataset) // args.train_batch_size
    max_iter = args.max_epoch * epoch_size

    ctc_loss = nn.CTCLoss(blank=len(CHARS) - 1, reduction='mean')  # reduction: 'none' | 'mean' | 'sum'
    if args.resume_epoch > 0:
        start_iter = args.resume_epoch * epoch_size
    else:
        start_iter = 0

    for iteration in range(start_iter, max_iter):
        if iteration % epoch_size == 0:
            # create batch iterator
            batch_iterator = iter(
                DataLoader(train_dataset, args.train_batch_size, shuffle=True, num_workers=args.num_workers,
                           collate_fn=collate_fn))
            loss_val = 0
            epoch += 1

        # if iteration !=0 and iteration % args.save_interval == 0:
        # torch.save(lprnet.state_dict(), args.save_folder + 'LPRNet_' + '_iteration_' + repr(iteration) + '.pth')

        # if (iteration + 1) % args.test_interval == 0:
        #     # Greedy_Decode_Eval(lprnet, test_dataset, args)
        #     acc = Greedy_Decode_Eval(lprnet, test_dataset, args)
        #     torch.save(lprnet.state_dict(), args.save_folder + 'LPRNet_' + '_accuracy_' + repr(acc) + '.pth')
        # lprnet.train() # should be switch to train mode

        start_time = time.time()
        # load train data
        images, labels, lengths = next(batch_iterator)
        # labels = np.array([el.numpy() for el in labels]).T
        # print(labels)
        # get ctc parameters
        input_lengths, target_lengths = sparse_tuple_for_ctc(T_length, lengths)
        # update lr
        lr = adjust_learning_rate(optimizer, epoch, args.learning_rate, args.lr_schedule)

        if args.cuda:
            images = Variable(images, requires_grad=False).cuda()
            labels = Variable(labels, requires_grad=False).cuda()
        else:
            images = Variable(images, requires_grad=False)
            labels = Variable(labels, requires_grad=False)

        # forward
        logits = lprnet(images)
        log_probs = logits.permute(2, 0, 1)  # for ctc loss: T x N x C
        # print(labels.shape)
        log_probs = log_probs.log_softmax(2).requires_grad_()
        # log_probs = log_probs.detach().requires_grad_()
        # print(log_probs.shape)
        # backprop
        optimizer.zero_grad()
        loss = ctc_loss(log_probs, labels, input_lengths=input_lengths, target_lengths=target_lengths)
        if loss.item() == np.inf:
            continue
        loss.backward()
        optimizer.step()
        loss_val += loss.item()
        end_time = time.time()
        if iteration % 20 == 0:
            print('Epoch:' + repr(epoch) + ' || epochiter: ' + repr(iteration % epoch_size) + '/' + repr(epoch_size)
                  + '|| Totel iter ' + repr(iteration) + ' || Loss: %.4f||' % (loss.item()) +
                  'Batch time: %.4f sec. ||' % (end_time - start_time) + 'LR: %.8f' % (lr))
        if iteration % epoch_size == 0:
            acc = Greedy_Decode_Eval(lprnet, test_dataset, args)
            value = [epoch, loss.item(), acc]
            write_excel(excel_name, sheet_name, value)
            if (acc > bestAcc):
                bestAcc = acc
                torch.save(lprnet.state_dict(), args.save_folder + 'best_' + '_accuracy_' + repr(acc) + '.pth')

    # final test
    print("Final test Accuracy:")
    Greedy_Decode_Eval(lprnet, test_dataset, args)

    # save final parameters
    torch.save(lprnet.state_dict(), args.save_folder + 'last_LPRNet_model.pth')


def Greedy_Decode_Eval(Net, datasets, args):
    # TestNet = Net.eval()
    epoch_size = len(datasets) // args.test_batch_size
    batch_iterator = iter(
        DataLoader(datasets, args.test_batch_size, shuffle=True, num_workers=args.num_workers, collate_fn=collate_fn))

    Tp = 0
    Tn_1 = 0
    Tn_2 = 0
    t1 = time.time()
    for i in range(epoch_size):
        # load train data
        images, labels, lengths = next(batch_iterator)
        start = 0
        targets = []
        for length in lengths:
            label = labels[start:start + length]
            targets.append(label)
            start += length
        targets = np.array([el.numpy() for el in targets], dtype=object)
        imgs = images.numpy().copy()

        if args.cuda:
            images = Variable(images.cuda())
        else:
            images = Variable(images)

        # forward
        # images: [bs, 3, 24, 94]
        # prebs:  [bs, 68, 18]
        prebs = Net(images)
        # greedy decode
        prebs = prebs.cpu().detach().numpy()
        preb_labels = list()
        for i in range(prebs.shape[0]):
            preb = prebs[i, :, :]  # 对每张图片 [68, 18]
            preb_label = list()
            for j in range(preb.shape[1]):  # 18  返回序列中每个位置最大的概率对应的字符idx  其中'-'是67
                preb_label.append(np.argmax(preb[:, j], axis=0))
            no_repeat_blank_label = list()
            pre_c = preb_label[0]
            if pre_c != len(CHARS) - 1:  # 记录重复字符
                no_repeat_blank_label.append(pre_c)
            for c in preb_label:  # 去除重复字符和空白字符'-'
                if (pre_c == c) or (c == len(CHARS) - 1):
                    if c == len(CHARS) - 1:
                        pre_c = c
                    continue
                no_repeat_blank_label.append(c)
                pre_c = c
            preb_labels.append(no_repeat_blank_label)  # 得到最终的无重复字符和无空白字符的序列
        for i, label in enumerate(preb_labels):  # 统计准确率
            # show image and its predict label
            if args.show:
                print(1)
                # show(imgs[i], label, targets[i])
            if len(label) != len(targets[i]):
                Tn_1 += 1  # 错误+1
                continue
            if (np.asarray(targets[i]) == np.asarray(label)).all():
                Tp += 1  # 完全正确+1
            else:
                Tn_2 += 1
    Acc = Tp * 1.0 / (Tp + Tn_1 + Tn_2)
    print("[Info] Test Accuracy: {} [{}:{}:{}:{}]".format(Acc, Tp, Tn_1, Tn_2, (Tp + Tn_1 + Tn_2)))
    t2 = time.time()
    print("[Info] Test Speed: {}s 1/{}]".format((t2 - t1) / len(datasets), len(datasets)))
    return round(Acc, 5)


if __name__ == "__main__":
    train()

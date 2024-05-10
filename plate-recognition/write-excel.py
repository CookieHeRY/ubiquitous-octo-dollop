import os.path
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

def write_excel(excel_name, sheet_name, value):
    columns = ["epoc", "train_loss","test_acc"]  # 列名
    # 创建一个 pandas 的数据框
    if not os.path.exists(excel_name):  # 文件不存在,就创建一个
        df = pd.DataFrame(columns=columns)
        df.to_excel(excel_name, index=False)
    books = load_workbook(excel_name).sheetnames  # 得到已存在的sheet列表

    # 将数据框写入 Excel 文件
    if sheet_name not in books:  # 如果sheet_name不存在,创建
        with pd.ExcelWriter(excel_name, engine="openpyxl", mode="a") as writer:
            df = pd.DataFrame(columns=columns)
            df.to_excel(writer, sheet_name=sheet_name, index=False)  # header=None 参数用于追加写入时不重复写入列名
    # 追加一行数据
    workbooks = load_workbook(excel_name)  # 获取文件
    worksheet = workbooks[sheet_name]  # 获取工作表sheet
    worksheet._current_row = worksheet.max_row  # 指定最大行数
    worksheet.append(value)  # 添加数据
    workbooks.save(excel_name)  # 保存

def image_show(excel_name, sheet_name):
    dataframe = pd.read_excel(excel_name, sheet_name)  # 读取excel的地址
    epoc = dataframe['epoc'].values  # 读出某列的数据为一个列表
    train_loss = dataframe['train_loss'].values  # 读出某列的数据为一个列表
    # train_acc = dataframe['train_acc'].values  # 读出某列的数据为一个列表
    test_acc = dataframe['test_acc'].values  # 读出某列的数据为一个列表
    plt.figure(1, dpi=300)  # 画第一个图
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 显示中文
    # plt.xlim([-0.1, 210])
    plt.ylim([-0.1, 3.5])  # 设置y轴刻度范围
    # plt.yticks(np.linspace(0, 2, 10))  # 设置刻度
    # plt.xticks(range(0, 21, 2))  # 共20个值，每2个点显示一次
    plt.plot(epoc, train_loss, color='r', marker='', linewidth=1, linestyle='-', label="loss")  # 可以调整粗细，大小，颜色
    plt.title("train_loss曲线")  # 显示图名称
    plt.xlabel('epoc')  # 显示x轴名称
    plt.ylabel('loss')  # 显示y轴名称
    plt.legend()  # 显示标签
    # -------------------------------
    plt.figure(2, dpi=300)  # 画第二个图
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 显示中文
    plt.ylim([0, 1])
    plt.yticks(np.linspace(0, 1, 10))
    # plt.plot(epoc, train_acc, color='r', marker='', linewidth=1, linestyle='--', label="train_acc")
    plt.plot(epoc, test_acc, color='b', marker='', linewidth=1, linestyle='-', label="test_acc")
    plt.title("accuracy准确率")
    plt.xlabel('epoc')
    plt.ylabel('accuracy')
    plt.legend()
    plt.show()  # 显示

# #测试
excel_name = r"results-210epoch.xlsx"
sheet_name = "Sheet1"  # 工作簿sheet的名字
image_show(excel_name, sheet_name)
#测试
# excel_name = r"example_pandas.xlsx"
# sheet_name = "Sheet1"
# value = [1, 2, 3, 4]
# write_excel(excel_name, sheet_name, value)
